import openpyxl

class HomePageData:
    @staticmethod
    def get_test_data(test_case_name):
        dict1 = {}
        final_data = []
        book = openpyxl.load_workbook("../Uitility/People data.xlsx")
        sheet = book.active
        test_cases = [cell.value for cell in sheet['A']]
        if test_case_name == 'all':
            for k in range(1, len(test_cases)):
                for i in range(1, sheet.max_row + 1):
                    if sheet.cell(row=i, column=1).value == test_cases[k]:
                        for j in range(2, sheet.max_column + 1):
                            dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                        final_data.append(dict1.copy())
            return final_data
        else:
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == test_case_name:
                    for j in range(2, sheet.max_column + 1):
                        dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    final_data.append(dict1)
                    print(final_data)
            return final_data





